"""
Excel report generation for SWIFT CSCF assessments.

This module uses the CSCF Assessment Excel template as a base and
fills per-control status (in-place / not-in-place / not-applicable)
into the appropriate worksheet cells.

NOTE: The default cell mapping here is intentionally conservative and
meant as a starting point. You should adjust CONTROL_SHEET_CONFIG to
match the exact layout of your template (sheet names and cell addresses).
"""

from io import BytesIO
from pathlib import Path
from typing import Dict, Any, Optional
import logging

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import MergedCell

from backend.config import settings
from backend.models.schemas import (
    SwiftExcelReportRequest,
    SwiftControlStatus,
    SwiftUserBackgroundData,
)


TEMPLATE_FILENAME = "CSCF_Assessment_Template_for_Mandatory_Controls_v2025_2.0.xlsx"


def get_template_path() -> Path:
    """
    Resolve the path to the base Excel template.

    The template currently lives at the inner repo root:
        /root/sai/SentraIQ-main/SentraIQ-main/CSCF_Assessment_Template_for_Mandatory_Controls_v2025_2.0.xlsx

    Given settings.BASE_PATH == <repo>/SentraIQ-main/SentraIQ-main,
    we look in BASE_PATH first, then one level up.
    """
    # By default, look next to the backend package, i.e. at BASE_PATH
    template_path = settings.BASE_PATH / TEMPLATE_FILENAME

    # If not found there, also try one level up (outer project root)
    if not template_path.exists():
        fallback = settings.BASE_PATH.parent / TEMPLATE_FILENAME
        return fallback

    return template_path


def _find_guideline_row(ws) -> Optional[int]:
    """
    Find the row index on the given worksheet where column A contains
    the word 'Guideline'. Returns None if not found.
    """
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        value = cell.value
        if isinstance(value, str) and "Guideline" in value:
            return cell.row
    return None


def _populate_user_background_sheet(
    wb, user_background: Optional[SwiftUserBackgroundData]
) -> None:
    """
    Locate the 'User Background Data Sheet' worksheet (by header text) and
    populate the key cells with values from the request.

    The implementation is resilient to minor layout changes by:
    - Searching for the sheet that contains the title 'USER BACKGROUND DATA SHEET'
    - Matching left-hand labels in column A to known keys and writing into column B
    """
    if not user_background:
        return

    header_text = "USER BACKGROUND DATA SHEET"
    target_ws = None

    # Try to find the worksheet that contains the header text in the first few rows
    for ws in wb.worksheets:
        try:
            for row in ws.iter_rows(min_row=1, max_row=5, max_col=5):
                for cell in row:
                    if isinstance(cell.value, str) and header_text.lower() in cell.value.lower():
                        target_ws = ws
                        break
                if target_ws:
                    break
        except Exception:
            # If anything goes wrong scanning this sheet, skip it – we don't want to
            # fail the whole Excel generation just because of one worksheet.
            continue
        if target_ws:
            break

    if not target_ws:
        logging.warning(
            "Could not locate 'User Background Data Sheet' worksheet in CSCF template"
        )
        return

    # Map the visible label text in column A to attributes on SwiftUserBackgroundData
    label_to_attr: Dict[str, str] = {
        "Customer Name": "customer_name",
        "BIC": "bic",
        "Architecture Type": "architecture_type",
        "Assessment Start Date": "assessment_start_date",
        "Assessment End Date": "assessment_end_date",
        "CSCF Version": "cscf_version",
        "Assessor Firm": "assessor_firm",
        "Lead Assessor Name": "lead_assessor_name",
        "Lead Assessor Title": "lead_assessor_title",
        "Assessor Name(s)": "assessor_names",
    }

    # Iterate down column A, and when we see a recognised label, set the value in column B
    for row in target_ws.iter_rows(min_row=1, max_col=2):
        label_cell = row[0]
        value_cell = row[1] if len(row) > 1 else None

        if not isinstance(label_cell.value, str):
            continue

        label = label_cell.value.strip()
        attr_name = label_to_attr.get(label)
        if not attr_name or value_cell is None:
            continue

        try:
            value = getattr(user_background, attr_name, None)
        except Exception:
            value = None

        if value is None:
            continue

        previous = value_cell.value
        value_cell.value = value
        logging.debug(
            "Updated User Background sheet cell %s%d from %r to %r",
            value_cell.column_letter,
            value_cell.row,
            previous,
            value,
        )


def generate_cscf_excel(request: SwiftExcelReportRequest) -> BytesIO:
    """
    Generate an Excel workbook for a SWIFT CSCF assessment.

    Returns an in-memory BytesIO stream containing the .xlsx file.
    """
    template_path = get_template_path()
    if not template_path.exists():
        raise FileNotFoundError(
            f"CSCF template not found at {template_path}. "
            "Please place the CSCF_Assessment_Template_for_Mandatory_Controls_v2025_2.0.xlsx "
            "file at the repository root or update get_template_path()."
        )

    # NOTE: We never modify the template file on disk. We always
    # load it read-only, apply changes in-memory, and return a BytesIO.
    # So filesystem write permissions are NOT required for the template.
    # Load workbook - use default settings to preserve structure and data validation
    # data_only=False preserves formulas, which is important for the template structure
    wb = load_workbook(filename=str(template_path), keep_vba=False)

    # First, try to populate the User Background Data Sheet, if the caller
    # provided any user background information.
    try:
        # If architecture_type is not explicitly set in user_background, default
        # to the swift_architecture_type from the request (if available).
        user_bg = request.user_background
        if user_bg and not user_bg.architecture_type and request.swift_architecture_type:
            user_bg.architecture_type = request.swift_architecture_type
        _populate_user_background_sheet(wb, user_bg)
    except Exception as e:
        logging.warning(f"Failed to populate User Background Data Sheet: {e}")

    # Map internal status to Excel dropdown values
    status_to_excel_value: Dict[str, str] = {
        "in-place": "Yes",
        "not-in-place": "No",
        "not-applicable": "N/A",
    }

    # Iterate over provided control statuses and write into corresponding sheets
    for control in request.control_statuses:
        sheet_name = control.control_id

        if sheet_name not in wb.sheetnames:
            # Skip controls that don't have a corresponding worksheet in the template
            continue

        ws = wb[sheet_name]

        try:
            guideline_row = _find_guideline_row(ws)
            if not guideline_row:
                # If we can't find the Guideline row, skip this sheet rather than failing
                continue

            # In the template, the row containing the word "Guideline" in column A
            # is a header row. The next row typically contains explanatory text,
            # and the actual Yes/No dropdown for the question is on the row
            # *after* that. Empirically this is guideline_row + 2.
            target_row = guideline_row + 2
            target_cell_ref = f"D{target_row}"
            excel_value = status_to_excel_value.get(control.status)

            if excel_value is None:
                # Unknown status value; skip
                continue

            # Get the target cell. Some rows in this template are merged, and
            # writing to a MergedCell raises an error, so we need to ensure we
            # select a real (unmerged) cell in column D.
            cell = ws[target_cell_ref]

            # If we accidentally landed on a merged cell, scan downward a few
            # rows until we find the first real cell in column D.
            if isinstance(cell, MergedCell):
                found_real_cell = False
                for offset in range(1, 6):  # look up to 5 rows below
                    candidate_row = target_row + offset
                    candidate_ref = f"D{candidate_row}"
                    candidate_cell = ws[candidate_ref]
                    if not isinstance(candidate_cell, MergedCell):
                        target_row = candidate_row
                        target_cell_ref = candidate_ref
                        cell = candidate_cell
                        found_real_cell = True
                        break

                if not found_real_cell:
                    logging.warning(
                        f"Could not find editable D cell for control {control.control_id} "
                        f"in sheet {sheet_name} starting from row {target_row}"
                    )
                    continue

            # Now set the value directly – it must exactly match dropdown options:
            # "Yes", "No", or "N/A".
            original_value = cell.value
            cell.value = excel_value

            # Also write reasoning / evidence summary (if provided) into the
            # merged cell directly under the question.
            #
            # Layout (per your description and template):
            # - Question text: merged cells B{target_row}:C{target_row}
            # - Yes/No dropdown: D{target_row}
            # - Reasoning cell: merged cells B{target_row+1}:D{target_row+1}
            #
            # We only need to write to the top-left cell of the merged region,
            # which is B{target_row+1}.
            if control.answer_summary:
                reason_row = target_row + 1
                reason_cell_ref = f"B{reason_row}"
                try:
                    reason_cell = ws[reason_cell_ref]
                    # Even if this is part of a merged range, openpyxl treats
                    # the top-left cell as a normal writable Cell.
                    prev_reason_value = reason_cell.value
                    reason_cell.value = control.answer_summary

                    logging.debug(
                        f"Updated {sheet_name}!{reason_cell_ref} from "
                        f"{prev_reason_value!r} to {control.answer_summary!r}"
                    )
                except Exception as reason_err:
                    logging.warning(
                        f"Failed to write reasoning for control {control.control_id} "
                        f"in sheet {sheet_name} at {reason_cell_ref}: {reason_err}"
                    )

            # Optional debug logging
            logging.debug(
                f"Updated {sheet_name}!{target_cell_ref} from {original_value!r} to {excel_value!r}"
            )
            
        except Exception as e:
            # Log the error for debugging but don't fail the entire report
            import logging
            logging.warning(f"Failed to update cell {target_cell_ref} in sheet {sheet_name}: {str(e)}")
            continue

    # --- Debug: create a simple test sheet to verify Excel is modified ---
    # This confirms that changes made by openpyxl are actually persisted
    # in the file returned to the UI.
    from datetime import datetime
    debug_sheet_name = "SentraIQ_Test"
    if debug_sheet_name in wb.sheetnames:
        ws_debug = wb[debug_sheet_name]
    else:
        ws_debug = wb.create_sheet(title=debug_sheet_name)

    ws_debug["A1"] = "SentraIQ Excel Write Test"
    ws_debug["A2"] = f"Generated at: {datetime.utcnow().isoformat()}Z"
    ws_debug["A3"] = f"Controls in request: {len(request.control_statuses)}"

    # Save workbook to in-memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

